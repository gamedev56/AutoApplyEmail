import pandas as pd
from filelock import FileLock
from openpyxl import load_workbook
from bs4 import BeautifulSoup

file_path = "companies.xlsx"
lock_path = "companies.xlsx.lock"

from dotenv import load_dotenv
import os

load_dotenv(override=True)
file_path = os.getenv("SAVED_COMPANIES_EXCEL_PATH")
lock_path = file_path + ".lock"

def readFromExcelSheet(file_path, start_row, end_row):
    df = pd.read_csv(file_path)
    range_df = df.iloc[start_row:end_row]
    sampled_df = range_df.sample(n=len(range_df), random_state=42)
    
    return sampled_df

def update_url(name, url):
    lock = FileLock(lock_path)
    with lock:
        wb = load_workbook(file_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2): 
            if row[0].value == name:  
                row[1].value = url  
                break
        else:
            sheet.append([name, url])

        wb.save(file_path)
        wb.close()

def update_email(url, emails):
    lock = FileLock(lock_path)
    with lock:
        wb = load_workbook(file_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2): 
            if row[1].value == url: 
                email_column_index = 2  
                if emails == "No Email Found":
                    row[email_column_index].value = emails
                    break
                else:
                    row[email_column_index].value = ", ".join(emails)
                    break
        wb.save(file_path)
        wb.close()

def update_email_sent(name, email_sent):
    lock = FileLock(lock_path)
    with lock:
        wb = load_workbook(file_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2):  
            if row[0].value == name:  
                row[4].value = email_sent 
                break
        else:
            raise ValueError(f"Name '{name}' not found in the sheet.")

        wb.save(file_path)
        wb.close()

def saveHTML(response, url = "https://www.google.com" ,path=""):
    file_name = url.replace("https://", "").replace("/", "_") + f"{path}.html"
    
    if isinstance(response, BeautifulSoup):
        html_content = response.prettify()
    else:
        html_content = response

    with open(file_name, "w", encoding="utf-8") as f:
        f.write(html_content)
    
    print(f"Saved HTML content to {file_name}")

